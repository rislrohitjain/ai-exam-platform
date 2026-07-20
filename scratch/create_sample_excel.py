import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def create_sample_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exam Import Template"

    # Styling definitions
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo header
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    headers = [
        "category_name",
        "paper_code",
        "paper_title",
        "paper_description",
        "question_type",
        "question_content",
        "question_answer_key",
        "question_marks"
    ]
    
    ws.append(headers)
    
    # Format headers
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    sample_rows = [
        [
            "Software Engineering",
            "SE-2026-T-002",
            "Clean Architecture Deep Dive",
            "Advanced assessment covering SOLID principles and software clean layers.",
            "objective",
            "Which of the following is NOT one of the SOLID design principles?\nA) Single Responsibility Principle\nB) Open/Closed Principle\nC) Interface Segregation Principle\nD) Object Relational Principle",
            "D",
            2.0
        ],
        [
            "Software Engineering",
            "SE-2026-T-002",
            "Clean Architecture Deep Dive",
            "Advanced assessment covering SOLID principles and software clean layers.",
            "subjective",
            "Explain the Dependency Inversion Principle (DIP) and provide a conceptual example of how to implement it in code to decouple business logic from a database layer.",
            "High-level modules should not depend on low-level modules; both should depend on abstractions. Abstractions should not depend on details; details should depend on abstractions. Example: Instead of business logic creating a concrete PostgresRepository class directly, declare an abstract IRepository interface, make the business logic receive it via Constructor Injection, and let the concrete PostgresRepository implement IRepository.",
            8.0
        ],
        [
            "Web Development",
            "WD-2026-T-002",
            "CSS Layouts and Flexbox",
            "Fundamentals of modern CSS layout engines.",
            "objective",
            "Which CSS Flexbox property specifies the alignment of items along the main axis?\nA) align-items\nB) justify-content\nC) align-content\nD) flex-direction",
            "B",
            1.5
        ]
    ]

    for row in sample_rows:
        ws.append(row)

    # Style rows
    for row_idx in range(2, len(sample_rows) + 2):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = thin_border
            if col_idx in [5, 8]: # type, marks
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Set column widths
    column_widths = {
        1: 22, # category_name
        2: 18, # paper_code
        3: 28, # paper_title
        4: 35, # paper_description
        5: 16, # question_type
        6: 50, # question_content
        7: 45, # question_answer_key
        8: 15  # question_marks
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
    ws.row_dimensions[1].height = 28 # Header row height
    for r in range(2, len(sample_rows) + 2):
        ws.row_dimensions[r].height = 45 # Sample rows height

    output_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../app/static'))
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "sample_import.xlsx")
    wb.save(output_path)
    print(f"Sample Excel template saved successfully at: {output_path}")

if __name__ == "__main__":
    create_sample_excel()
