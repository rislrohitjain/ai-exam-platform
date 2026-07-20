from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Dict, Optional, Any
import io
import datetime
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
import logging

from app.core.database import get_db
from app.models.db_models import Category, QuestionPaper, Question, PlatformSettings, Certificate, ExamSubmission, User, Organization
from app.services.llm_factory import generate_embedding, get_settings
from app.services.pdf_service import verify_signature

from app.api.auth import require_admin, require_instructor, require_candidate

logger = logging.getLogger("admin_api")
router = APIRouter(prefix="/api/admin", tags=["Admin API"])

# --- Pydantic Schemas ---
class CategoryCreate(BaseModel):
    name: str = Field(..., max_length=255)
    parent_id: Optional[int] = None
    organization_id: Optional[int] = None

class CategoryUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=255)
    parent_id: Optional[int] = None
    organization_id: Optional[int] = None

class PaperCreate(BaseModel):
    code: Optional[str] = Field(None, max_length=50)
    title: str = Field(..., max_length=255)
    category_id: int
    organization_id: Optional[int] = None
    grade_thresholds: Dict[str, float] = Field(default_factory=lambda: {"A++": 90.0, "A": 80.0, "B": 70.0, "C": 50.0})
    description: Optional[str] = None

class UserUpdateRole(BaseModel):
    roles: str = Field(..., max_length=255)  # comma-separated roles, e.g. "instructor,candidate"
    organization_id: Optional[int] = None
    password: Optional[str] = Field(None, min_length=6, max_length=100)

class QuestionCreate(BaseModel):
    type: str = Field(..., pattern="^(objective|subjective)$")
    content: str
    answer_key: str
    marks: float = Field(..., gt=0)
    option_a: Optional[str] = None
    option_b: Optional[str] = None
    option_c: Optional[str] = None
    option_d: Optional[str] = None

class QuestionUpdate(BaseModel):
    type: Optional[str] = Field(None, pattern="^(objective|subjective)$")
    content: Optional[str] = None
    answer_key: Optional[str] = None
    marks: Optional[float] = Field(None, gt=0)
    option_a: Optional[str] = None
    option_b: Optional[str] = None
    option_c: Optional[str] = None
    option_d: Optional[str] = None

class SettingsUpdate(BaseModel):
    active_provider: str = Field(..., pattern="^(ollama|openai|groq|openrouter)$")
    ollama_base_url: Optional[str] = "http://localhost:11434"
    ollama_model: Optional[str] = "llama3"
    ollama_temperature: Optional[float] = 0.2
    openai_api_key: Optional[str] = ""
    openai_model: Optional[str] = "gpt-4o"
    openai_temperature: Optional[float] = 0.2
    groq_api_key: Optional[str] = ""
    groq_model: Optional[str] = "llama3-8b-8192"
    groq_temperature: Optional[float] = 0.2
    openrouter_api_key: Optional[str] = ""
    openrouter_model: Optional[str] = "meta-llama/llama-3-8b-instruct:free"
    openrouter_temperature: Optional[float] = 0.2

class CertificateVerifyRequest(BaseModel):
    certificate_id: str
    student_id: str
    paper_id: int
    signature: str

class OrganizationCreate(BaseModel):
    name: str = Field(..., max_length=255)
    type: str = Field("institute", max_length=100)
    description: Optional[str] = None

class OrganizationUpdate(BaseModel):
    name: Optional[str] = Field(None, max_length=255)
    type: Optional[str] = Field(None, max_length=100)
    description: Optional[str] = None

# --- Organization CRUD API (Admin only) ---
@router.post("/organizations", response_model=Dict[str, Any])
def create_organization(payload: OrganizationCreate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    existing = db.query(Organization).filter(Organization.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Organization name is already registered")
    
    org = Organization(
        name=payload.name,
        type=payload.type,
        description=payload.description
    )
    db.add(org)
    db.commit()
    db.refresh(org)
    return {"message": "Organization created successfully", "id": org.id, "name": org.name}

@router.get("/organizations", response_model=List[Dict[str, Any]])
def list_organizations(db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    orgs = db.query(Organization).all()
    return [{"id": o.id, "name": o.name, "type": o.type, "description": o.description} for o in orgs]

@router.put("/organizations/{org_id}", response_model=Dict[str, Any])
def update_organization(org_id: int, payload: OrganizationUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    org = db.query(Organization).filter(Organization.id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
        
    if payload.name is not None:
        org.name = payload.name
    if payload.type is not None:
        org.type = payload.type
    if payload.description is not None:
        org.description = payload.description
        
    db.commit()
    return {"message": "Organization updated successfully"}

@router.delete("/organizations/{org_id}", response_model=Dict[str, Any])
def delete_organization(org_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    org = db.query(Organization).filter(Organization.id == org_id).first()
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
        
    db.delete(org)
    db.commit()
    return {"message": "Organization deleted successfully"}

# --- Category API ---
@router.post("/categories", response_model=Dict[str, Any])
def create_category(payload: CategoryCreate, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    if payload.parent_id:
        parent = db.query(Category).filter(Category.id == payload.parent_id).first()
        if not parent:
            raise HTTPException(status_code=404, detail="Parent category not found")
            
    existing = db.query(Category).filter(Category.name == payload.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Category with this name already exists")
        
    org_id = payload.organization_id
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles:
        # Instructors can only create categories for their own organization
        org_id = current_user.organization_id

    category = Category(name=payload.name, parent_id=payload.parent_id, organization_id=org_id)
    db.add(category)
    db.commit()
    db.refresh(category)
    return {"message": "Category created successfully", "category_id": category.id, "name": category.name}

@router.get("/categories", response_model=List[Dict[str, Any]])
def get_categories_tree(db: Session = Depends(get_db), current_user: User = Depends(require_candidate)):
    """
    Returns the categories tree, filtered by user's organization tenancy if not system admin.
    """
    allowed_roles = current_user.roles.split(",")
    if "admin" in allowed_roles:
        categories = db.query(Category).all()
    else:
        categories = db.query(Category).filter(
            or_(
                Category.organization_id == current_user.organization_id,
                Category.organization_id == None
            )
        ).all()

    # Format tree structure
    nodes = {cat.id: {"id": cat.id, "name": cat.name, "parent_id": cat.parent_id, "children": []} for cat in categories}
    root_nodes = []
    
    for cat in categories:
        node = nodes[cat.id]
        if cat.parent_id is None:
            root_nodes.append(node)
        else:
            parent_node = nodes.get(cat.parent_id)
            if parent_node:
                parent_node["children"].append(node)
            else:
                root_nodes.append(node) # Fallback
                
    return root_nodes

@router.put("/categories/{category_id}", response_model=Dict[str, Any])
def update_category(category_id: int, payload: CategoryUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    cat = db.query(Category).filter(Category.id == category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
        
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles and cat.organization_id != current_user.organization_id:
        raise HTTPException(status_code=403, detail="Not authorized to edit this category")
        
    if payload.name is not None:
        existing = db.query(Category).filter(Category.name == payload.name, Category.id != category_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="Category with this name already exists")
        cat.name = payload.name
        
    if payload.parent_id is not None:
        if payload.parent_id == category_id:
            raise HTTPException(status_code=400, detail="Category cannot be its own parent")
        parent = db.query(Category).filter(Category.id == payload.parent_id).first()
        if not parent:
            raise HTTPException(status_code=404, detail="Parent category not found")
        cat.parent_id = payload.parent_id
    elif "parent_id" in payload.model_fields_set and payload.parent_id is None:
        cat.parent_id = None
        
    if "admin" in allowed_roles and "organization_id" in payload.model_fields_set:
        cat.organization_id = payload.organization_id
        
    db.commit()
    return {"message": "Category updated successfully"}

@router.delete("/categories/{category_id}", response_model=Dict[str, Any])
def delete_category(category_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    cat = db.query(Category).filter(Category.id == category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
        
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles and cat.organization_id != current_user.organization_id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this category")
        
    db.delete(cat)
    db.commit()
    return {"message": "Category deleted successfully"}

# --- Question Paper API ---
@router.post("/papers", response_model=Dict[str, Any])
def create_paper(payload: PaperCreate, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    # Validate category
    cat = db.query(Category).filter(Category.id == payload.category_id).first()
    if not cat:
        raise HTTPException(status_code=404, detail="Category not found")
        
    org_id = payload.organization_id
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles:
        # Instructors can only create papers under their own organization
        org_id = current_user.organization_id

    # Auto-generate unique paper code if not provided
    if not payload.code:
        prefix = "".join([w[0].upper() for w in cat.name.split() if w])[:2]
        if len(prefix) < 2:
            prefix = (cat.name[:2] + "XX").upper()[:2]
            
        year = datetime.datetime.now().year
        count = db.query(QuestionPaper).count() + 1
        payload.code = f"{prefix}-{year}-M-{count:03d}"
        
    existing = db.query(QuestionPaper).filter(QuestionPaper.code == payload.code).first()
    if existing:
        raise HTTPException(status_code=400, detail="Paper with this code already exists")
        
    paper = QuestionPaper(
        code=payload.code,
        title=payload.title,
        category_id=payload.category_id,
        organization_id=org_id,
        description=payload.description,
        grade_thresholds=payload.grade_thresholds
    )
    db.add(paper)
    db.commit()
    db.refresh(paper)
    return {"message": "Question paper created successfully", "paper_id": paper.id, "code": paper.code}

@router.get("/papers", response_model=List[Dict[str, Any]])
def list_papers(db: Session = Depends(get_db), current_user: User = Depends(require_candidate)):
    """
    Returns exam papers. Filtered by organization if user is not system admin.
    """
    allowed_roles = current_user.roles.split(",")
    if "admin" in allowed_roles:
        papers = db.query(QuestionPaper).all()
    else:
        papers = db.query(QuestionPaper).filter(
            or_(
                QuestionPaper.organization_id == current_user.organization_id,
                QuestionPaper.organization_id == None
            )
        ).all()

    return [{
        "id": p.id,
        "code": p.code,
        "title": p.title,
        "category_id": p.category_id,
        "total_marks": p.total_marks,
        "grade_thresholds": p.grade_thresholds,
        "description": p.description,
        "questions_count": len(p.questions),
        "organization_name": p.organization.name if p.organization else "System Global"
    } for p in papers]

@router.post("/papers/{paper_id}/questions", response_model=Dict[str, Any])
async def add_question_to_paper(paper_id: int, payload: QuestionCreate, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    paper = db.query(QuestionPaper).filter(QuestionPaper.id == paper_id).first()
    if not paper:
        raise HTTPException(status_code=404, detail="Question paper not found")
        
    # Security: Instructors can only modify papers of their own organization
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles and paper.organization_id != current_user.organization_id:
         raise HTTPException(status_code=403, detail="Not authorized to edit this exam paper")

    # Generate Embedding for semantic reference context of subjective questions
    embedding = None
    if payload.type == "subjective":
        context_text = f"{payload.content} {payload.answer_key}"
        try:
            embedding = await generate_embedding(db, context_text)
        except Exception as e:
            logger.error(f"Error creating embedding: {e}")
            
    question = Question(
        paper_id=paper_id,
        type=payload.type,
        content=payload.content,
        answer_key=payload.answer_key,
        marks=payload.marks,
        context_vector=embedding,
        option_a=payload.option_a,
        option_b=payload.option_b,
        option_c=payload.option_c,
        option_d=payload.option_d
    )
    db.add(question)
    
    # Update paper total marks
    paper.total_marks += payload.marks
    db.commit()
    db.refresh(question)
    
    return {
        "message": "Question added to paper successfully",
        "question_id": question.id,
        "marks": question.marks,
        "paper_total_marks": paper.total_marks,
        "vector_generated": embedding is not None
    }

@router.get("/papers/{paper_id}/questions", response_model=List[Dict[str, Any]])
def list_questions(paper_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_candidate)):
    paper = db.query(QuestionPaper).filter(QuestionPaper.id == paper_id).first()
    if not paper:
        raise HTTPException(status_code=404, detail="Question paper not found")
        
    # Attempt lock: Candidates can only get the questions if they haven't submitted this paper yet
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles and "instructor" not in allowed_roles:
        from app.models.db_models import ExamSubmission
        existing_submission = db.query(ExamSubmission).filter(
            ExamSubmission.student_id == current_user.username,
            ExamSubmission.paper_id == paper_id
        ).first()
        if existing_submission:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST, 
                detail="You have already attempted this exam. Only one attempt is allowed."
            )
            
    return [{
        "id": q.id,
        "type": q.type,
        "content": q.content,
        "marks": q.marks,
        "answer_key": q.answer_key,
        "has_embedding": q.context_vector is not None,
        "option_a": q.option_a,
        "option_b": q.option_b,
        "option_c": q.option_c,
        "option_d": q.option_d
    } for q in paper.questions]

# --- Settings API (Admin only) ---
@router.get("/settings", response_model=Dict[str, Any])
def get_platform_settings(db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    s = get_settings(db)
    return {
        "active_provider": s.active_provider,
        "ollama_base_url": s.ollama_base_url,
        "ollama_model": s.ollama_model,
        "ollama_temperature": s.ollama_temperature,
        "openai_model": s.openai_model,
        "openai_temperature": s.openai_temperature,
        "openai_api_key_configured": bool(s.openai_api_key),
        "groq_model": s.groq_model,
        "groq_temperature": s.groq_temperature,
        "groq_api_key_configured": bool(s.groq_api_key),
        "openrouter_model": s.openrouter_model,
        "openrouter_temperature": s.openrouter_temperature,
        "openrouter_api_key_configured": bool(s.openrouter_api_key)
    }

@router.put("/settings", response_model=Dict[str, Any])
def update_platform_settings(payload: SettingsUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    s = get_settings(db)
    s.active_provider = payload.active_provider
    if payload.ollama_base_url:
        s.ollama_base_url = payload.ollama_base_url
    if payload.ollama_model:
        s.ollama_model = payload.ollama_model
    if payload.ollama_temperature is not None:
        s.ollama_temperature = payload.ollama_temperature
    if payload.openai_api_key is not None:
        s.openai_api_key = payload.openai_api_key
    if payload.openai_model:
        s.openai_model = payload.openai_model
    if payload.openai_temperature is not None:
        s.openai_temperature = payload.openai_temperature
    if payload.groq_api_key is not None:
        s.groq_api_key = payload.groq_api_key
    if payload.groq_model:
        s.groq_model = payload.groq_model
    if payload.groq_temperature is not None:
        s.groq_temperature = payload.groq_temperature
    if payload.openrouter_api_key is not None:
        s.openrouter_api_key = payload.openrouter_api_key
    if payload.openrouter_model:
        s.openrouter_model = payload.openrouter_model
    if payload.openrouter_temperature is not None:
        s.openrouter_temperature = payload.openrouter_temperature
        
    db.commit()
    logger.info("Platform settings updated dynamically.")
    return {"message": "Settings updated dynamically", "active_provider": s.active_provider}

# --- Certificates Dashboard (Admin only) ---
@router.get("/certificates", response_model=List[Dict[str, Any]])
def list_certificates(db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    certs = db.query(Certificate).all()
    return [{
        "id": c.id,
        "student_id": c.student_id,
        "paper_id": c.paper_id,
        "paper_code": c.paper.code if c.paper else "N/A",
        "paper_title": c.paper.title if c.paper else "N/A",
        "issue_date": c.issue_date.strftime("%Y-%m-%d %H:%M:%S"),
        "signature": c.digital_signature
    } for c in certs]

@router.post("/certificates/verify", response_model=Dict[str, Any])
def verify_certificate_authenticity(payload: CertificateVerifyRequest, current_user: User = Depends(require_admin)):
    """
    Cryptographically verifies the authenticity of a certificate signature.
    """
    is_valid = verify_signature(
        student_id=payload.student_id,
        paper_id=payload.paper_id,
        cert_id=payload.certificate_id,
        signature=payload.signature
    )
    if is_valid:
        return {"status": "valid", "message": "Cryptographic signature matches database payload and platform key."}
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid certificate signature. Verification failed."
        )

# --- Analytics API ---
@router.get("/dashboard", response_model=Dict[str, Any])
def get_analytics_dashboard(db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    """
    Returns dashboard analytics. Filtered by organization if user is not system admin.
    """
    allowed_roles = current_user.roles.split(",")
    if "admin" in allowed_roles:
        submissions = db.query(ExamSubmission).filter(ExamSubmission.status == "evaluated").all()
    else:
        submissions = db.query(ExamSubmission).join(QuestionPaper).filter(
            ExamSubmission.status == "evaluated",
            QuestionPaper.organization_id == current_user.organization_id
        ).all()
    
    total_submissions = len(submissions)
    if total_submissions == 0:
        return {
            "total_submissions": 0,
            "average_score": 0.0,
            "pass_fail_ratio": {"pass": 0, "fail": 0},
            "grade_distribution": {},
            "performance_by_paper": []
        }
        
    grades_count = {}
    pass_count = 0
    fail_count = 0
    total_score = 0.0
    paper_scores = {}
    
    for sub in submissions:
        grade = sub.final_grade or "F"
        grades_count[grade] = grades_count.get(grade, 0) + 1
        
        if sub.percentage and sub.percentage >= 50.0:
            pass_count += 1
        else:
            fail_count += 1
            
        total_score += sub.overall_score or 0.0
        
        p_id = sub.paper_id
        if p_id not in paper_scores:
            paper_scores[p_id] = {"title": sub.paper.title if sub.paper else "N/A", "scores": [], "code": sub.paper.code if sub.paper else "N/A"}
        paper_scores[p_id]["scores"].append(sub.percentage or 0.0)

    perf_by_paper = []
    for pid, data in paper_scores.items():
        avg_pct = sum(data["scores"]) / len(data["scores"])
        perf_by_paper.append({
            "paper_id": pid,
            "code": data["code"],
            "title": data["title"],
            "average_percentage": avg_pct,
            "submissions_count": len(data["scores"])
        })

    return {
        "total_submissions": total_submissions,
        "average_score": total_score / total_submissions,
        "pass_fail_ratio": {"pass": pass_count, "fail": fail_count},
        "grade_distribution": grades_count,
        "performance_by_paper": perf_by_paper
    }

# --- User Management API (Admin only) ---
@router.get("/users", response_model=List[Dict[str, Any]])
def list_users(db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    users = db.query(User).all()
    return [{
        "id": u.id, 
        "username": u.username, 
        "plain_password": u.plain_password or "********",
        "roles": u.roles,
        "organization_id": u.organization_id,
        "organization_name": u.organization.name if u.organization else "System Global"
    } for u in users]

@router.put("/users/{username}/role", response_model=Dict[str, Any])
def update_user_role(username: str, payload: UserUpdateRole, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
        
    allowed_roles = payload.roles.split(",")
    if user.username == "admin" and "admin" not in allowed_roles:
        raise HTTPException(status_code=400, detail="Cannot downgrade primary admin account")
        
    user.roles = payload.roles
    user.organization_id = payload.organization_id
    
    if payload.password:
        from app.api.auth import hash_password
        user.hashed_password = hash_password(payload.password)
        user.plain_password = payload.password
        logger.info(f"Password reset triggered for user {username}")
        
    db.commit()
    return {"message": "User roles, organization, and password updated successfully"}

@router.delete("/users/{username}", response_model=Dict[str, Any])
def delete_user(username: str, db: Session = Depends(get_db), current_user: User = Depends(require_admin)):
    user = db.query(User).filter(User.username == username).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    if user.username == "admin":
        raise HTTPException(status_code=400, detail="Cannot delete primary admin account")
    db.delete(user)
    db.commit()
    return {"message": "User deleted successfully"}

# --- Question Management CRUD ---
@router.put("/questions/{question_id}", response_model=Dict[str, Any])
async def update_question(question_id: int, payload: QuestionUpdate, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    q = db.query(Question).filter(Question.id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
        
    # Security: Instructors can only modify questions of their own organization
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles and q.paper and q.paper.organization_id != current_user.organization_id:
         raise HTTPException(status_code=403, detail="Not authorized to edit questions on this exam paper")

    old_marks = q.marks
    if payload.type is not None:
        q.type = payload.type
    if payload.content is not None:
        q.content = payload.content
    if payload.answer_key is not None:
        q.answer_key = payload.answer_key
        
    if payload.option_a is not None:
        q.option_a = payload.option_a
    if payload.option_b is not None:
        q.option_b = payload.option_b
    if payload.option_c is not None:
        q.option_c = payload.option_c
    if payload.option_d is not None:
        q.option_d = payload.option_d
        
    if payload.marks is not None:
        q.marks = payload.marks
        if q.paper:
            q.paper.total_marks = q.paper.total_marks - old_marks + payload.marks
            
    # Re-calculate embedding for vector search if subjective grading contexts change
    if q.type == "subjective" and (payload.content is not None or payload.answer_key is not None):
        context_text = f"{q.content} {q.answer_key}"
        try:
            q.context_vector = await generate_embedding(db, context_text)
        except Exception as e:
            logger.error(f"Error updating embedding context: {e}")
            
    db.commit()
    return {"message": "Question updated successfully"}

@router.delete("/questions/{question_id}", response_model=Dict[str, Any])
def delete_question(question_id: int, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    q = db.query(Question).filter(Question.id == question_id).first()
    if not q:
        raise HTTPException(status_code=404, detail="Question not found")
        
    # Security: Instructors can only modify questions of their own organization
    allowed_roles = current_user.roles.split(",")
    if "admin" not in allowed_roles and q.paper and q.paper.organization_id != current_user.organization_id:
         raise HTTPException(status_code=403, detail="Not authorized to edit questions on this exam paper")

    if q.paper:
        q.paper.total_marks -= q.marks
        
    db.delete(q)
    db.commit()
    return {"message": "Question deleted successfully"}

# --- Excel Exporter API ---
@router.get("/export/{data_type}")
def export_data(data_type: str, db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    wb = Workbook()
    ws = wb.active
    ws.title = data_type.capitalize()
    
    allowed_roles = current_user.roles.split(",")
    is_admin = "admin" in allowed_roles
    org_id = current_user.organization_id
    
    if data_type == "users":
        if not is_admin:
            raise HTTPException(status_code=403, detail="Only admins can export user accounts")
        ws.append(["Username", "Roles", "Organization"])
        users = db.query(User).all()
        for u in users:
            org_name = u.organization.name if u.organization else "System Global"
            ws.append([u.username, u.roles, org_name])
            
    elif data_type == "categories":
        ws.append(["Category ID", "Category Name", "Parent ID", "Organization"])
        if is_admin:
            cats = db.query(Category).all()
        else:
            cats = db.query(Category).filter(Category.organization_id == org_id).all()
        for c in cats:
            org_name = c.organization.name if c.organization else "System Global"
            ws.append([c.id, c.name, c.parent_id, org_name])
            
    elif data_type == "papers":
        ws.append(["Paper Code", "Title", "Category Name", "Total Marks", "Description", "Organization"])
        if is_admin:
            papers = db.query(QuestionPaper).all()
        else:
            papers = db.query(QuestionPaper).filter(QuestionPaper.organization_id == org_id).all()
        for p in papers:
            cat_name = p.category.name if p.category else "N/A"
            org_name = p.organization.name if p.organization else "System Global"
            ws.append([p.code, p.title, cat_name, p.total_marks, p.description, org_name])
            
    elif data_type == "questions":
        ws.append(["Question ID", "Paper Code", "Paper Title", "Type", "Content", "Option A", "Option B", "Option C", "Option D", "Answer Key / Rubric", "Marks"])
        if is_admin:
            questions = db.query(Question).all()
        else:
            questions = db.query(Question).join(QuestionPaper).filter(QuestionPaper.organization_id == org_id).all()
        for q in questions:
            paper_code = q.paper.code if q.paper else "N/A"
            paper_title = q.paper.title if q.paper else "N/A"
            ws.append([q.id, paper_code, paper_title, q.type, q.content, q.option_a, q.option_b, q.option_c, q.option_d, q.answer_key, q.marks])
            
    elif data_type == "submissions":
        ws.append(["Submission ID", "Student ID", "Paper Code", "Paper Title", "Status", "Score Obtained", "Max Marks", "Percentage", "Grade", "Submission Date"])
        if is_admin:
            subs = db.query(ExamSubmission).all()
        else:
            subs = db.query(ExamSubmission).join(QuestionPaper).filter(QuestionPaper.organization_id == org_id).all()
        for s in subs:
            paper_code = s.paper.code if s.paper else "N/A"
            paper_title = s.paper.title if s.paper else "N/A"
            ws.append([
                s.id, s.student_id, paper_code, paper_title, s.status, 
                s.overall_score or 0.0, 
                sum(q.marks for q in s.paper.questions) if (s.paper and s.paper.questions) else 0.0,
                s.percentage or 0.0, s.final_grade or "N/A",
                s.created_at.strftime("%Y-%m-%d %H:%M:%S") if s.created_at else "N/A"
            ])
    else:
        raise HTTPException(status_code=400, detail="Invalid export data type")
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"export_{data_type}.xlsx"
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- Excel Importer API ---
@router.post("/import/papers")
async def import_papers_excel(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: User = Depends(require_instructor)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported.")
        
    try:
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file: {e}")
        
    headers = [cell.value for cell in ws[1]]
    required = ["category_name", "paper_code", "paper_title", "question_type", "question_content", "question_answer_key", "question_marks"]
    for col in required:
        if col not in headers:
            raise HTTPException(status_code=400, detail=f"Missing required column: '{col}'")
            
    h_map = {name: headers.index(name) for name in headers if name}
    
    allowed_roles = current_user.roles.split(",")
    is_admin = "admin" in allowed_roles
    org_id = current_user.organization_id

    imported_papers = {}
    imported_questions_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
            
        cat_name = str(row[h_map["category_name"]]).strip()
        p_code = str(row[h_map["paper_code"]]).strip()
        p_title = str(row[h_map["paper_title"]]).strip()
        p_desc = str(row[h_map.get("paper_description", -1)]) if "paper_description" in h_map else ""
        if p_desc == "None" or p_desc == "-1":
            p_desc = ""
            
        q_type = str(row[h_map["question_type"]]).strip().lower()
        q_content = str(row[h_map["question_content"]]).strip()
        q_key = str(row[h_map["question_answer_key"]]).strip()
        
        try:
            q_marks = float(row[h_map["question_marks"]])
        except Exception:
            q_marks = 1.0
            
        if not cat_name or not p_code or not p_title or not q_content or not q_key:
            continue
            
        if q_type not in ["objective", "subjective"]:
            q_type = "objective"
            
        # 1. Resolve Category (scoped by organization)
        if is_admin:
            cat = db.query(Category).filter(Category.name == cat_name).first()
        else:
            cat = db.query(Category).filter(Category.name == cat_name, Category.organization_id == org_id).first()
            
        if not cat:
            cat = Category(name=cat_name, organization_id=org_id)
            db.add(cat)
            db.commit()
            db.refresh(cat)
            
        # 2. Resolve Question Paper
        paper = db.query(QuestionPaper).filter(QuestionPaper.code == p_code).first()
        if not paper:
            paper = QuestionPaper(
                code=p_code,
                title=p_title,
                category_id=cat.id,
                organization_id=org_id,
                description=p_desc,
                grade_thresholds={"A": 80, "B": 60, "C": 50}
            )
            db.add(paper)
            db.commit()
            db.refresh(paper)
        else:
            # Check tenancy authorization for existing paper
            if not is_admin and paper.organization_id != org_id:
                continue # Skip papers from other organizations
            
        imported_papers[paper.id] = paper
        
        # 3. Create Question
        embedding = None
        if q_type == "subjective":
            context_text = f"{q_content} {q_key}"
            try:
                embedding = await generate_embedding(db, context_text)
            except Exception as emb_err:
                logger.error(f"Error generating embedding: {emb_err}")
                
        q_opt_a = str(row[h_map["option_a"]]).strip() if "option_a" in h_map and row[h_map["option_a"]] is not None else None
        q_opt_b = str(row[h_map["option_b"]]).strip() if "option_b" in h_map and row[h_map["option_b"]] is not None else None
        q_opt_c = str(row[h_map["option_c"]]).strip() if "option_c" in h_map and row[h_map["option_c"]] is not None else None
        q_opt_d = str(row[h_map["option_d"]]).strip() if "option_d" in h_map and row[h_map["option_d"]] is not None else None

        question = Question(
            paper_id=paper.id,
            type=q_type,
            content=q_content,
            answer_key=q_key,
            marks=q_marks,
            context_vector=embedding,
            option_a=q_opt_a,
            option_b=q_opt_b,
            option_c=q_opt_c,
            option_d=q_opt_d
        )
        db.add(question)
        imported_questions_count += 1
        
    db.commit()
    
    # 4. Recalculate total marks
    for pid, p in imported_papers.items():
        total = sum(q.marks for q in p.questions)
        p.total_marks = total
    db.commit()
    
    return {
        "message": "Data imported successfully.",
        "imported_papers_count": len(imported_papers),
        "imported_questions_count": imported_questions_count
    }
